"""
XLSX extractor for DOCINTEL.
Extracts text and metadata from XLSX files using openpyxl.
"""

from pathlib import Path
from typing import Optional

import openpyxl

from docintel.common.exceptions import ExtractionError
from docintel.common.logging import get_logger
from docintel.common.constants import SOURCE_LAYER_NATIVE

logger = get_logger(__name__)


class XLSXExtractor:
    """
    Extracts text and metadata from XLSX files.
    Uses openpyxl for reading Excel spreadsheets.
    """
    
    def __init__(self):
        """Initialize the XLSX extractor."""
        pass
    
    def extract(self, file_path: str) -> dict:
        """
        Extract text and metadata from an XLSX file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Dictionary with 'text', 'metadata', and 'source_layer'
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Extract metadata
            metadata = self._extract_metadata(wb)
            
            # Extract text from all sheets
            text_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = self._extract_sheet_text(sheet, sheet_name)
                if sheet_text:
                    text_parts.append(f"\n[Sheet: {sheet_name}]\n{sheet_text}")
            
            wb.close()
            
            full_text = "\n".join(text_parts)
            
            logger.info(
                "XLSX extraction complete",
                extra={
                    "file": file_path,
                    "sheets": len(wb.sheetnames),
                    "total_chars": len(full_text),
                },
            )
            
            return {
                "text": full_text,
                "metadata": metadata,
                "source_layer": SOURCE_LAYER_NATIVE,
            }
            
        except Exception as e:
            error_msg = f"Failed to extract XLSX: {str(e)}"
            logger.error(error_msg, extra={"file": file_path}, exc_info=True)
            raise ExtractionError(error_msg) from e
    
    def _extract_sheet_text(self, sheet, sheet_name: str) -> str:
        """
        Extract text from a worksheet.
        
        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            Extracted text
        """
        try:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to strings
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    rows.append(" | ".join(cells))
            return "\n".join(rows)
        except Exception as e:
            logger.warning(f"Could not extract sheet {sheet_name}: {str(e)}")
            return ""
    
    def _extract_metadata(self, wb: openpyxl.Workbook) -> dict:
        """
        Extract metadata from an XLSX workbook.
        
        Args:
            wb: openpyxl workbook object
            
        Returns:
            Dictionary of metadata
        """
        metadata = {}
        
        try:
            # Get workbook properties
            props = wb.properties
            metadata["title"] = props.title or ""
            metadata["subject"] = props.subject or ""
            metadata["creator"] = props.creator or ""
            metadata["created"] = str(props.created) if props.created else ""
            metadata["modified"] = str(props.modified) if props.modified else ""
            
            # Get sheet names
            metadata["sheet_names"] = wb.sheetnames
            metadata["sheet_count"] = len(wb.sheetnames)
            
            # Get active sheet
            metadata["active_sheet"] = wb.active.title if wb.active else ""
            
        except Exception as e:
            logger.warning(f"Could not extract all XLSX metadata: {str(e)}")
        
        return metadata
    
    def extract_sheet(self, file_path: str, sheet_name: str) -> dict:
        """
        Extract text from a specific sheet.
        
        Args:
            file_path: Path to the XLSX file
            sheet_name: Name of the sheet to extract
            
        Returns:
            Dictionary with 'text' and 'metadata'
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            
            sheet = wb[sheet_name]
            text = self._extract_sheet_text(sheet, sheet_name)
            
            metadata = {
                "sheet_name": sheet_name,
                "source_format": "XLSX",
            }
            
            wb.close()
            
            logger.info(
                "XLSX sheet extraction complete",
                extra={"file": file_path, "sheet": sheet_name, "chars": len(text)},
            )
            
            return {
                "text": text,
                "metadata": metadata,
                "source_layer": SOURCE_LAYER_NATIVE,
            }
            
        except Exception as e:
            error_msg = f"Failed to extract XLSX sheet: {str(e)}"
            logger.error(error_msg, extra={"file": file_path}, exc_info=True)
            raise ExtractionError(error_msg) from e
    
    def get_sheet_names(self, file_path: str) -> list[str]:
        """
        Get all sheet names in an XLSX file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            List of sheet names
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            error_msg = f"Failed to get sheet names: {str(e)}"
            logger.error(error_msg, extra={"file": file_path})
            raise ExtractionError(error_msg) from e


# Global XLSX extractor instance
_xlsx_extractor: Optional[XLSXExtractor] = None


def get_xlsx_extractor() -> XLSXExtractor:
    """Get the global XLSX extractor instance."""
    global _xlsx_extractor
    if _xlsx_extractor is None:
        _xlsx_extractor = XLSXExtractor()
    return _xlsx_extractor